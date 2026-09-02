"""Tests voor de Synergrid-profielenparser.

RLP0N en SLP-EX zijn `.xlsb`, een formaat dat niet met openpyxl geschreven
kan worden — deze tests schrijven daarom kleine `.xlsx`-fixtures met
openpyxl en lezen ze met `engine="openpyxl"`. `_find_sheet` en de
kolomlogica zijn engine-onafhankelijk (pyxlsb en openpyxl geven allebei een
`pandas.ExcelFile`/DataFrame terug), dus dat dekt dezelfde code als in
productie loopt tegen de echte `.xlsb`-bestanden — enkel het bestandsformaat
van de fixture verschilt. De volledige keten is bovendien manueel
geverifieerd tegen de echte, gedownloade Synergrid-bestanden (zie de
pipeline_report.json-structuur en de commit die deze module invoert).
"""

from __future__ import annotations

import openpyxl
import pytest

from energie_vlaanderen.ingest.profielen.workbook import (
    ProfielenWorkbookError,
    ProfielenWorkbookParser,
)


def _schrijf_nationaal_werkboek(pad, sheet_naam="ENU_UTC"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_naam
    ws.append(["UTC", "Year", "Month", "Day", "h", "Min", "From", "To", "ENU", "BinL", "BinH"])
    ws.append(["2027-01-01T00:00:00", 2027, 1, 1, 0, 0, "", "", 0.5, 1, 0])
    ws.append(["2027-01-01T00:15:00", 2027, 1, 1, 0, 15, "", "", 0.5, 1, 0])
    wb.save(pad)


def _schrijf_breed_werkboek(pad, sheet_naam="RLP96UbyDGO"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_naam
    # 3 headerrijen: groep, netbeheerdernaam, GLN — kolommen 0..5 zijn
    # tijd-id's, kolom 6 is de overbodige datumserie, waardekolommen vanaf 7.
    ws.append([None] * 7 + ["RLP", "GROEP2"])
    ws.append([None] * 7 + ["Testbeheerder Een", "Testbeheerder Twee"])
    ws.append(["CET", "Year", "Month", "Day", "h", "Min", "Date", "1234567890123", "9876543210987"])
    ws.append(["2027-01-01T00:00:00", 2027, 1, 1, 0, 0, "", 0.1, 0.2])
    ws.append(["2027-01-01T00:15:00", 2027, 1, 1, 0, 15, "", 0.1, 0.2])
    wb.save(pad)


class TestNationaalProfiel:
    def test_leest_tijdstip_en_waarde(self, tmp_path):
        pad = tmp_path / "slp_ex.xlsx"
        _schrijf_nationaal_werkboek(pad)

        parser = ProfielenWorkbookParser()
        result = parser.parse_nationaal(pad, engine="openpyxl", sheet_bevat="utc", waarde_kolom="ENU")

        assert len(result.rows) == 2
        # Expliciete UTC-suffix, ook zonder CET-verschuiving: een naïeve
        # string zonder tijdzone-aanduiding zou PostgreSQL via de
        # sessietijdzone laten interpreteren i.p.v. als UTC.
        assert result.rows[0].tijdstip == "2027-01-01T00:00:00+00:00"
        assert result.rows[0].waarde == 0.5
        assert result.rows[0].netbeheerder_gln is None

    def test_gebruikt_laatste_kolom_als_geen_waardekolom_opgegeven(self, tmp_path):
        pad = tmp_path / "spp.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SPP_ex-ante_2027"
        ws.append(["UTC", "Year", "Month", "Day", "Hour", "Min", "SPPExanteBE"])
        ws.append(["2027-01-01T00:00:00", 2027, 1, 1, 0, 0, 0.075])
        wb.save(pad)

        parser = ProfielenWorkbookParser()
        result = parser.parse_nationaal(pad, engine="openpyxl", sheet_bevat="ex-ante")

        assert result.rows[0].waarde == 0.075

    def test_onbestaande_sheet_faalt_hard(self, tmp_path):
        # Meerdere sheets, geen ervan bevat de zoekterm: `_find_sheet` mag
        # hier niet gokken. (Bij precies één sheet in het werkboek wordt die
        # altijd gebruikt, ongeacht de naam — zo zijn de echte RLP0N/GOS-
        # bestanden opgebouwd. Dat pad wordt door de andere tests gedekt.)
        pad = tmp_path / "leeg.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "IetsAnders"
        wb.create_sheet("NogIetsAnders")
        wb.save(pad)

        parser = ProfielenWorkbookParser()
        with pytest.raises(ProfielenWorkbookError):
            parser.parse_nationaal(pad, engine="openpyxl", sheet_bevat="utc")


class TestBreedProfiel:
    def test_meldt_naar_lange_vorm_per_netbeheerder(self, tmp_path):
        pad = tmp_path / "rlp0n.xlsx"
        _schrijf_breed_werkboek(pad)

        parser = ProfielenWorkbookParser()
        result = parser.parse_breed_per_netbeheerder(pad, engine="openpyxl", sheet_bevat="bydgo")

        # 2 tijdstippen x 2 netbeheerders
        assert len(result.rows) == 4
        glns = {row.netbeheerder_gln for row in result.rows}
        assert glns == {"1234567890123", "9876543210987"}
        namen = {row.netbeheerder_gln: row.netbeheerder_naam for row in result.rows}
        assert namen["1234567890123"] == "Testbeheerder Een"

    def test_geen_gln_kolommen_faalt_hard(self, tmp_path):
        pad = tmp_path / "geen_gln.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RLP96UbyDGO"
        ws.append([None] * 7 + ["RLP"])
        ws.append([None] * 7 + ["Iemand"])
        ws.append(["CET", "Year", "Month", "Day", "h", "Min", "Date", "GeenGLN"])
        ws.append(["2027-01-01T00:00:00", 2027, 1, 1, 0, 0, "", 0.1])
        wb.save(pad)

        parser = ProfielenWorkbookParser()
        with pytest.raises(ProfielenWorkbookError):
            parser.parse_breed_per_netbeheerder(pad, engine="openpyxl", sheet_bevat="bydgo")
